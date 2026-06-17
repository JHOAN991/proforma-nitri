"""
Generador de Proformas - Nutricereales de Panamá, S.A.
Streamlit app que llena la plantilla Nitri01.xlsx preservando formato.

Uso:
    streamlit run proforma_app.py
"""
from __future__ import annotations

import re
import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

# ---------------------------------------------------------------------------
# Configuración de rutas
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
TEMPLATE_FILE = BASE_DIR / "Nitri01.xlsx"
OUTPUT_DIR = BASE_DIR / "proformas_generadas"
OUTPUT_DIR.mkdir(exist_ok=True)

MAX_LINEAS = 46              # filas 16..61
FILA_INICIO = 16
FILA_FIN = 61
COL_TOTAL_FORMULA = 'K'      # ya tiene =IF(J="","",I*J*B)

EMPTY_ITEM = {
    "codigo": "",
    "descripcion": "",
    "um": 0,
    "cantidad": 0,
    "precio": 0.0,
}

# ---------------------------------------------------------------------------
# Carga del catálogo (Hoja2) con cache
# ---------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def cargar_catalogo(path: Path, _mtime: float) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Hoja2", dtype={"No. Producto": str})
    df.columns = [str(c).strip() for c in df.columns]
    # Renombrar la columna "tamaño" sin importar el encoding
    for col in df.columns:
        if col.lower().startswith("tama"):
            df = df.rename(columns={col: "tamano"})
            break
    df["No. Producto"] = df["No. Producto"].astype(str).str.strip()
    df["Descripcion Producto"] = df["Descripcion Producto"].astype(str).str.strip()
    df["Categoria"] = df["Categoria"].astype(str).str.strip()
    df["UM"] = df["tamano"].apply(extraer_um)
    return df


def extraer_um(texto: str) -> int:
    """Extrae el primer número de 'XX/YYY GRAMOS' → XX (unidades por caja)."""
    if not texto:
        return 0
    m = re.match(r"\s*(\d+)\s*/", str(texto))
    return int(m.group(1)) if m else 0


def leer_numero_proforma(path: Path) -> int:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        valor = wb["Hoja1"]["K3"].value
    finally:
        wb.close()
    try:
        return int(valor) if valor is not None else 1
    except (TypeError, ValueError):
        return 1


# ---------------------------------------------------------------------------
# Generación del archivo Excel
# ---------------------------------------------------------------------------
def generar_excel(
    template: Path,
    destino: Path,
    numero: int,
    fecha: date,
    fecha_exp: date | None,
    cliente: dict,
    entrega: dict,
    contacto: dict,
    descuento: float,
    items: list[dict],
) -> None:
    shutil.copy(template, destino)
    wb = openpyxl.load_workbook(destino)
    ws = wb["Hoja1"]

    # Encabezado
    ws["K3"] = int(numero)
    ws["K4"] = fecha
    if fecha_exp:
        ws["K5"] = fecha_exp

    # Cliente
    ws["A10"] = cliente.get("nombre", "")
    ws["A11"] = cliente.get("ruc", "")
    ws["A12"] = cliente.get("direccion1", "")
    ws["A13"] = cliente.get("direccion2", "")
    ws["A14"] = cliente.get("telefono", "")

    # Entrega
    ws["E10"] = entrega.get("direccion", "")
    ws["E14"] = entrega.get("pais", "")

    # Contacto
    ws["I10"] = contacto.get("correo", "")
    ws["I11"] = contacto.get("contacto", "")

    # Limpiar líneas previas (A..J; la K es fórmula y se conserva)
    for fila in range(FILA_INICIO, FILA_FIN + 1):
        for col in ("A", "B", "C", "I", "J"):
            ws[f"{col}{fila}"] = None

    # Escribir items
    for idx, item in enumerate(items):
        fila = FILA_INICIO + idx
        if fila > FILA_FIN:
            break
        ws[f"A{fila}"] = item["codigo"]
        ws[f"B{fila}"] = int(item["um"]) if item["um"] else None
        ws[f"C{fila}"] = item["descripcion"]
        ws[f"I{fila}"] = float(item["cantidad"])
        ws[f"J{fila}"] = float(item["precio"])

    # Descuento
    ws["K68"] = float(descuento or 0)

    wb.save(destino)
    wb.close()


def actualizar_numero_plantilla(template: Path, nuevo_numero: int) -> None:
    """Persiste el siguiente número de proforma en la plantilla (K3)."""
    wb = openpyxl.load_workbook(template)
    wb["Hoja1"]["K3"] = int(nuevo_numero)
    wb.save(template)
    wb.close()


# ---------------------------------------------------------------------------
# UI - Streamlit
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Proformas Nutricereales", page_icon="📋", layout="wide")
st.title("📋 Generador de Proformas")
st.caption("NUTRICEREALES DE PANAMÁ, S.A.")

if not TEMPLATE_FILE.exists():
    st.error(f"No se encontró la plantilla en: {TEMPLATE_FILE}")
    st.stop()

catalogo = cargar_catalogo(TEMPLATE_FILE, TEMPLATE_FILE.stat().st_mtime)

# Estado de sesión
if "items" not in st.session_state:
    st.session_state["items"] = []
if "numero" not in st.session_state:
    st.session_state["numero"] = leer_numero_proforma(TEMPLATE_FILE)

# ---------------------------------------------------------------------------
# Sidebar: encabezado y cliente
# ---------------------------------------------------------------------------
with st.sidebar:
    st.header("📝 Datos de la proforma")
    numero = st.number_input(
        "Número de proforma",
        min_value=1,
        value=int(st.session_state["numero"]),
        step=1,
    )
    col_f1, col_f2 = st.columns(2)
    with col_f1:
        fecha = st.date_input("Fecha", value=date.today(), format="DD/MM/YYYY")
    with col_f2:
        fecha_exp = st.date_input(
            "Expira", value=None, format="DD/MM/YYYY"
        )

    st.divider()
    st.subheader("👤 Cliente")
    cliente_nombre = st.text_input("Nombre del cliente")
    cliente_ruc = st.text_input("RUC")
    cliente_dir1 = st.text_input("Dirección línea 1")
    cliente_dir2 = st.text_input("Dirección línea 2 (ciudad/zona)")
    cliente_tel = st.text_input("Teléfono")

    with st.expander("🚚 Entrega (opcional)"):
        entrega_dir = st.text_area("Dirección de entrega", height=80)
        entrega_pais = st.text_input("País", value="PANAMÁ")

    with st.expander("✉️ Contacto (opcional)"):
        contacto_correo = st.text_input("Correo electrónico")
        contacto_nombre = st.text_input("Contacto")

    descuento = st.number_input("Descuento USD", min_value=0.0, value=0.0, step=1.0, format="%.2f")

# ---------------------------------------------------------------------------
# Sección principal: catálogo + items
# ---------------------------------------------------------------------------
col_busq, col_items = st.columns([1, 1])

with col_busq:
    st.subheader("🔍 Buscar producto")

    categorias = ["(Todas)"] + sorted(catalogo["Categoria"].unique().tolist())
    cat_sel = st.selectbox("Categoría", categorias)
    query = st.text_input("Código o nombre (parcial)", placeholder="ej: IM001, spaghetti")

    df = catalogo.copy()
    if cat_sel != "(Todas)":
        df = df[df["Categoria"] == cat_sel]
    if query:
        q = query.strip().lower()
        df = df[
            df["No. Producto"].str.lower().str.contains(q, na=False)
            | df["Descripcion Producto"].str.lower().str.contains(q, na=False)
        ]

    st.caption(f"{len(df)} productos")
    if df.empty:
        st.info("Sin resultados.")
    else:
        # Selector + agregar
        df_view = df[["No. Producto", "Descripcion Producto", "tamano", "UM"]].reset_index(drop=True)
        df_view.columns = ["Código", "Descripción", "Tamaño", "UM"]
        st.dataframe(df_view, use_container_width=True, hide_index=True, height=300)

        opciones = [
            f"{row['No. Producto']} — {row['Descripcion Producto']}"
            for _, row in df.iterrows()
        ]
        sel = st.selectbox("Seleccionar producto", opciones, key="sel_producto")
        codigo_sel = sel.split(" — ", 1)[0] if sel else ""

        c1, c2, c3 = st.columns([1, 1, 1])
        with c1:
            cantidad = st.number_input("Cantidad (cajas)", min_value=1, value=1, step=1)
        with c2:
            precio = st.number_input("Precio unitario", min_value=0.0, value=0.0, step=0.0001, format="%.4f")
        with c3:
            st.write("")
            st.write("")
            agregar = st.button("➕ Agregar", use_container_width=True, type="primary")

        if agregar:
            if len(st.session_state["items"]) >= MAX_LINEAS:
                st.error(f"Máximo {MAX_LINEAS} líneas por proforma.")
            elif precio <= 0:
                st.error("El precio debe ser mayor a 0.")
            else:
                row = catalogo[catalogo["No. Producto"] == codigo_sel].iloc[0]
                st.session_state["items"].append({
                    "codigo": row["No. Producto"],
                    "descripcion": row["Descripcion Producto"],
                    "um": int(row["UM"]),
                    "cantidad": float(cantidad),
                    "precio": float(precio),
                })
                st.success(f"Agregado: {codigo_sel}")
                st.rerun()

with col_items:
    st.subheader(f"🧾 Líneas en la proforma ({len(st.session_state['items'])}/{MAX_LINEAS})")

    if not st.session_state["items"]:
        st.info("No hay líneas todavía. Busca y agrega productos a la izquierda.")
    else:
        # Tabla editable de items
        df_items = pd.DataFrame(st.session_state["items"])
        df_items["total"] = df_items["cantidad"] * df_items["precio"] * df_items["um"]
        df_show = df_items.rename(columns={
            "codigo": "Código",
            "descripcion": "Descripción",
            "um": "UM",
            "cantidad": "Cant",
            "precio": "Precio Un",
            "total": "Total",
        })
        edited = st.data_editor(
            df_show,
            use_container_width=True,
            hide_index=False,
            disabled=["Código", "Descripción", "UM", "Total"],
            column_config={
                "Precio Un": st.column_config.NumberColumn(format="$%.4f"),
                "Total": st.column_config.NumberColumn(format="$%.2f"),
                "Cant": st.column_config.NumberColumn(format="%d"),
            },
            key="editor_items",
        )

        # Aplicar ediciones
        for i, row in edited.iterrows():
            st.session_state["items"][i]["cantidad"] = float(row["Cant"])
            st.session_state["items"][i]["precio"] = float(row["Precio Un"])

        # Botón borrar fila
        idx_borrar = st.number_input(
            "Eliminar línea #", min_value=0, max_value=len(st.session_state["items"]) - 1, value=0, step=1
        )
        if st.button("🗑️ Eliminar línea seleccionada"):
            st.session_state["items"].pop(int(idx_borrar))
            st.rerun()

        # Totales
        subtotal = sum(i["cantidad"] * i["precio"] * i["um"] for i in st.session_state["items"])
        total = subtotal - descuento
        st.markdown(
            f"""
            **Subtotal:** `${subtotal:,.2f}`
            **Descuento:** `${descuento:,.2f}`
            ### TOTAL USD: `${total:,.2f}`
            """
        )

st.divider()

# ---------------------------------------------------------------------------
# Acciones
# ---------------------------------------------------------------------------
col_a, col_b, col_c = st.columns([1, 1, 1])

with col_a:
    if st.button("🧹 Limpiar proforma", use_container_width=True):
        st.session_state["items"] = []
        st.rerun()

with col_b:
    auto_incrementar = st.checkbox("Auto-incrementar K3 al guardar", value=True)

with col_c:
    generar = st.button("💾 Generar proforma", use_container_width=True, type="primary")

if generar:
    if not cliente_nombre.strip():
        st.error("Falta el nombre del cliente.")
    elif not st.session_state["items"]:
        st.error("Agrega al menos una línea de producto.")
    else:
        cliente_safe = re.sub(r"[^A-Za-z0-9_-]+", "_", cliente_nombre)[:40].strip("_")
        nombre_out = f"Proforma_{int(numero):05d}_{cliente_safe}.xlsx"
        destino = OUTPUT_DIR / nombre_out
        try:
            generar_excel(
                template=TEMPLATE_FILE,
                destino=destino,
                numero=int(numero),
                fecha=fecha,
                fecha_exp=fecha_exp,
                cliente={
                    "nombre": cliente_nombre,
                    "ruc": cliente_ruc,
                    "direccion1": cliente_dir1,
                    "direccion2": cliente_dir2,
                    "telefono": cliente_tel,
                },
                entrega={"direccion": entrega_dir, "pais": entrega_pais},
                contacto={"correo": contacto_correo, "contacto": contacto_nombre},
                descuento=descuento,
                items=st.session_state["items"],
            )
            if auto_incrementar:
                actualizar_numero_plantilla(TEMPLATE_FILE, int(numero) + 1)
                st.session_state["numero"] = int(numero) + 1

            st.success(f"✅ Proforma generada: `{destino}`")
            with open(destino, "rb") as f:
                st.download_button(
                    "⬇️ Descargar archivo",
                    data=f.read(),
                    file_name=nombre_out,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"Error al generar: {e}")
            raise
